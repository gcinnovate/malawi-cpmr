import os
# import sys
import string
import random
from openpyxl import load_workbook
import click
from dotenv import load_dotenv
from flask_migrate import Migrate, upgrade
from app import create_app, db, redis_client
from app.models import (
    Location, LocationTree, PoliceStation, User, Role,
    FlowData, JusticeCourt, SummaryCases, TraditionalAuthority,
    CommunityVictimSupportUnit, ChildrensCorner, OneStopCenter)
import datetime
from flask import current_app
from sqlalchemy.sql import text
from getpass import getpass
from config import INDICATOR_NAME_MAPPING, INDICATORS_TO_SWAP_KEYVALS

dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)

COV = None
if os.environ.get('FLASK_COVERAGE'):
    import coverage
    COV = coverage.coverage(branch=True, include='app/*')
    COV.start()


app = create_app(os.getenv('FLASK_CONFIG') or 'default')
migrate = Migrate(app, db)


@app.before_first_request
def before_first_request_func():
    locs = Location.query.filter_by(level=3).all()
    districts = {}
    for l in locs:
        districts[l.name] = {'id': l.id, 'parent_id': l.parent_id}
    redis_client.districts = districts

    stations = PoliceStation.query.all()
    police_stations = {}
    for s in stations:
        police_stations[s.name] = s.id
    redis_client.police_stations = police_stations

    courts = JusticeCourt.query.all()
    justice_courts = {}
    for c in courts:
        justice_courts[c.name] = c.id
    redis_client.justice_courts = justice_courts

    print("This function will run once")


@app.shell_context_processor
def make_shell_context():
    return dict(app=app, db=db, User=User, Role=Role)


@app.teardown_appcontext
def teardown_db(exception=None):
    db.session.remove()


@app.cli.command("initdb")
def initdb():
    def id_generator(size=12, chars=string.ascii_lowercase + string.ascii_uppercase + string.digits):
        return ''.join(random.choice(chars) for _ in range(size))

    Role.insert_roles()
    country = Location.query.filter_by(name='Malawi', level=1).all()
    if country:
        click.echo("Database Already Initialized")
        return

    click.echo("Database Initialization Starting.............!")

    db.session.add(LocationTree(name='Malawi Administrative Divisions'))
    db.session.commit()
    # Add country
    db.session.add(Location(name='Malawi', code=id_generator(), tree_id=1))  # Country
    # Add the regions
    db.session.add_all(
        [
            Location(name='Central', code=id_generator(), parent_id=1, tree_id=1),  # 2
            Location(name='Eastern', code=id_generator(), parent_id=1, tree_id=1),  # 3
            Location(name='Northern', code=id_generator(), parent_id=1, tree_id=1),  # 4
            Location(name='Southern', code=id_generator(), parent_id=1, tree_id=1),  # 5
        ]
    )

    # Central = 2, Eastern = 3, Northern = 4, Southern = 5
    regions_data = {
        '2': [
            'Dedza', 'Dowa', 'Kasungu', 'Lilongwe', 'Mchinji',
            'Nkhotakota', 'Ntcheu', 'Ntchisi', 'Salima'],
        '3': ['Balaka', 'Machinga', 'Mangochi', 'Zomba'],
        '4': ['Chitipa', 'Karonga', 'Likoma', 'Mzimba', 'Nkhata Bay', 'Rumphi'],
        '5': [
            'Blantyre', 'Chikwawa', 'Chiradzulu', 'Mulanje', 'Mwanza', 'Neno', 'Nsanje',
            'Phalombe', 'Thyolo']
    }
    for k, v in regions_data.items():
        for val in v:
            db.session.add(Location(name=val, code=id_generator(), parent_id=k, tree_id=1))
    db.session.commit()

    # Each district name corresponds to a Police Station in Malawi
    districts = Location.query.filter_by(level=3).all()
    for d in districts:
        db.session.add(PoliceStation(name=d.name, district_id=d.id))
        db.session.add(JusticeCourt(name=d.name, district_id=d.id))

    blantyre = Location.query.filter_by(name='Blantyre', level=3).all()[0]
    db.session.add(PoliceStation(name='Chileka', district_id=blantyre.id))
    db.session.add(PoliceStation(name='Limbe', district_id=blantyre.id))

    mzimba = Location.query.filter_by(name='Mzimba', level=3).all()[0]
    db.session.add(PoliceStation(name='Mzuzu', district_id=mzimba.id))

    nkhotakota = Location.query.filter_by(name='Nkhotakota', level=3).all()[0]
    db.session.add(PoliceStation(name='Nkhunga', district_id=nkhotakota.id))

    dowa = Location.query.filter_by(name='Dowa', level=3).all()[0]
    db.session.add(PoliceStation(name='Mponela', district_id=dowa.id))

    lilongwe = Location.query.filter_by(name='Lilongwe', level=3).all()[0]
    db.session.add(PoliceStation(name='Kanengo', district_id=lilongwe.id))

    db.session.commit()
    click.echo("Database Initialization Complete!")


@app.cli.command("load_test_data")
@click.option('--report', '-r', default='pvsu')
@click.option('--start-year', '-s', default=2016)
@click.option('--end-year', '-e', default=2020)
@click.option('--start-month', '-m', default=1)
@click.option('--end-month', '-n', default=13)
@click.option('--init', '-i', default=0)
@click.option('--limit', '-x', default="yes")  # whether to limit on years to create data for
def load_test_data(report, start_year, end_year, start_month, end_month, init, limit):
    # for pvsu, diversion, cvsu and cc test data
    from config import INDICATOR_CATEGORY_MAPPING, INDICATOR_THRESHOLD
    # print(report)
    if report in ('pvsu', 'diversion'):
        stations = PoliceStation.query.all()
    elif report == 'cvsu':
        stations = CommunityVictimSupportUnit.query.all()
    elif report == 'cc':
        stations = ChildrensCorner.query.all()
    elif report == 'osc':
        stations = OneStopCenter.query.all()

    year = datetime.datetime.now().year
    if start_year == end_year:
        end_year += 1
    for y in range(start_year, end_year):
        for m in range(start_month, end_month):
            if y == year and m > datetime.datetime.now().month - 1:
                if limit == "yes":
                    continue
                else:
                    pass
            click.echo("{0}-{1:02}".format(y, m))
            for p in stations:
                district_id = p.district_id
                region_id = Location.query.filter_by(id=district_id).first().parent_id
                month_str = "{0}-{1:02}".format(y, m)
                total_cases = 0
                boys_total = 0
                girls_total = 0
                men_total = 0
                women_total = 0
                values = {}
                for k, v in INDICATOR_CATEGORY_MAPPING.get(report).items():
                    indcators_total = 0
                    for ind in v:
                        if k in INDICATORS_TO_SWAP_KEYVALS:
                            field = "{0}_{1}".format(k, ind)
                        else:
                            field = "{0}_{1}".format(ind, k)
                        if init:
                            val = 0
                        else:
                            val = random.choice(range(INDICATOR_THRESHOLD[k]))
                        values[field] = val
                        indcators_total += val
                        if ind == 'boys':
                            boys_total += val
                        elif ind == 'girls':
                            girls_total += val
                        elif ind == 'men':
                            men_total += val
                        elif ind == 'women':
                            women_total += val
                    values[k] = indcators_total
                    total_cases += indcators_total
                # values['total_cases'] = total_cases
                if report in ('pvsu', 'cvsu'):
                    values['total_cases'] = total_cases
                    values['boys_total'] = boys_total
                    values['girls_total'] = girls_total
                    values['men_total'] = men_total
                    values['women_total'] = women_total

                if report in ('pvsu', 'diversion'):  # reporting at police station level
                    flow_data_obj = FlowData.query.filter_by(
                        year=y, month=month_str, region=region_id, district=district_id,
                        station=p.id, report_type=report).first()
                    if not flow_data_obj:
                        db.session.add(FlowData(
                            region=region_id, district=district_id, station=p.id, month=month_str,
                            year=y, report_type=report, values=values))
                elif report == 'cvsu':
                    flow_data_obj = FlowData.query.filter_by(
                        year=y, month=month_str, region=region_id, district=district_id,
                        cvsu=p.id, report_type=report).first()
                    if not flow_data_obj:
                        db.session.add(FlowData(
                            region=region_id, district=district_id, cvsu=p.id, month=month_str,
                            year=y, report_type=report, values=values))
                elif report == 'cc':
                    flow_data_obj = FlowData.query.filter_by(
                        year=y, month=month_str, region=region_id, district=district_id,
                        children_corner=p.id, report_type=report).first()
                    if not flow_data_obj:
                        db.session.add(FlowData(
                            region=region_id, district=district_id, children_corner=p.id,
                            month=month_str, year=y, report_type=report, values=values))
                elif report == 'osc':
                    flow_data_obj = FlowData.query.filter_by(
                        year=y, month=month_str, region=region_id, district=district_id,
                        one_stop_center=p.id, report_type=report).first()
                    if not flow_data_obj:
                        db.session.add(FlowData(
                            region=region_id, district=district_id, one_stop_center=p.id,
                            month=month_str, year=y, report_type=report, values=values))

                click.echo(values)
            db.session.commit()


@app.cli.command("load_test_data2")
@click.option('--report', '-r', default='ncjf')
@click.option('--start-year', '-s', default=2016)
@click.option('--end-year', '-e', default=2020)
@click.option('--start-month', '-m', default=1)
@click.option('--end-month', '-n', default=13)
@click.option('--init', '-i', default=0)
@click.option('--limit', '-x', default="yes")  # whether to limit on years to create data for
def load_test_data2(report, start_year, end_year, start_month, end_month, init, limit):
    from config import INDICATOR_CATEGORY_MAPPING, INDICATOR_THRESHOLD
    print(report)
    justice_courts = JusticeCourt.query.all()
    year = datetime.datetime.now().year
    if start_year == end_year:
        end_year += 1
    for y in range(start_year, end_year):
        for m in range(start_month, end_month):
            if y == year and m > datetime.datetime.now().month - 1:
                if limit == "yes":
                    continue
                else:
                    pass
            click.echo("{0}-{1:02}".format(y, m))
            for p in justice_courts:
                district_id = p.district_id
                region_id = Location.query.filter_by(id=district_id).first().parent_id
                month_str = "{0}-{1:02}".format(y, m)
                total_cases = 0
                total_childcases = 0
                total_concluded = 0
                values = {}
                for k, v in INDICATOR_CATEGORY_MAPPING.get(report).items():
                    indcators_total = 0
                    if not v:  # if indicator has no sub categories!
                        field = "{0}".format(k)
                        if init:
                            val = 0
                        else:
                            val = random.choice(range(INDICATOR_THRESHOLD[k]))
                        values[field] = val
                        continue

                    for ind in v:
                        field = "{0}_{1}".format(ind, k)
                        if init:
                            val = 0
                        else:
                            val = random.choice(range(INDICATOR_THRESHOLD[k]))
                        values[field] = val
                        indcators_total += val
                        if ind in ('fromprevmonth', 'newlyregistered'):
                            total_childcases += val
                        if ind in ('concluded'):
                            total_concluded += val
                    values[k] = indcators_total
                    total_cases += indcators_total
                    values['total_childcases'] = total_childcases
                    values['concluded'] = total_concluded
                values['total_cases'] = total_cases
                flow_data_obj = FlowData.query.filter_by(
                    year=y, month=month_str, region=region_id, district=district_id,
                    court=p.id, report_type=report).first()
                if not flow_data_obj:  # only create entry if none is existing
                    db.session.add(FlowData(
                        region=region_id, district=district_id, court=p.id, month=month_str,
                        year=y, report_type=report, values=values))
                click.echo(values)
            db.session.commit()


@app.cli.command()
def deploy():
    """Run deployment tasks."""
    # migrate database to latest revision
    upgrade()


@app.cli.command("create-user")
def createuser():
    username = input("Enter Username: ")
    email = input("Enter Email: ")
    password = getpass()
    cpass = getpass("Confirm Password: ")
    assert password == cpass
    u = User(username=username, email=email)
    u.password = cpass
    db.session.add(u)
    db.session.commit()
    u.confirmed = True
    db.session.commit()
    click.echo("User added!")


@app.cli.command("create-views")
def create_views():
    with current_app.open_resource('../views.sql') as f:
        # print(f.read())
        click.echo("Gonna create views")
        db.session.execute(text(f.read().decode('utf8')))
        db.session.commit()
        click.echo("Done creating views")


@app.cli.command("refresh-pvsu-casetypes")
def refresh_pvsu_casetypes():
    results = db.engine.execute("SELECT * FROM pvsu_casetypes_regional_view order by year desc")
    # print(results.keys())
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3],
            region=r[4], report_type='pvsu', summary_for='region', summary_slug='types').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='pvsu', summary_for='region', summary_slug='types')
            db.session.add(s)
        db.session.commit()

    # Load data for pvsu regional demography
    results = db.engine.execute("SELECT * FROM pvsu_cases_demographics_regional_view order by year desc")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3], region=r[4],
            report_type='pvsu', summary_for='region', summary_slug='demography').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='pvsu', summary_for='region', summary_slug='demography')
            db.session.add(s)
        db.session.commit()

    results = db.engine.execute("SELECT * from ncjf_childvictim_categories_view ORDER BY year desc, month desc;")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        for k in results.keys():
            if k in ('month', 'year'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3],
            report_type='ncjf', summary_for='nation', summary_slug='childvictim_categories').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                report_type='ncjf', summary_for='nation', summary_slug='childvictim_categories')
            db.session.add(s)
        db.session.commit()

    results = db.engine.execute("SELECT * from ncjf_casetypes_national_view ORDER BY year desc, month desc;")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        casetypes = {
            'cvc': 'Child Victim', 'cbc':
            'Child Beneficaries', 'inconflict':
            'Children in conflict with the law'}
        values = {'cvc': {}, 'cbc': {}, 'inconflict': {}}
        for k in results.keys():
            if k in ('month', 'year'):
                continue
            ctype, vkey = k.split('_')
            values[vkey][ctype] = row[k]
        for key, val in values.items():
            records.append((casetypes[key], val, month, year))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=r[0], month=r[2], year=r[3],
            report_type='ncjf', summary_for='nation', summary_slug='child_cases').first()
        if summary:
            summary.json_values = r[1]
        else:
            s = SummaryCases(
                casetype=r[0], json_values=r[1], month=r[2], year=r[3],
                report_type='ncjf', summary_for='nation', summary_slug='child_cases')
            db.session.add(s)
        db.session.commit()

    results = db.engine.execute("SELECT * FROM cvsu_casetypes_regional_view order by year desc")
    # print(results.keys())
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3],
            region=r[4], report_type='cvsu', summary_for='region', summary_slug='types').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='cvsu', summary_for='region', summary_slug='types')
            db.session.add(s)
        db.session.commit()

    # Load data for cvsu regional demography
    results = db.engine.execute("SELECT * FROM cvsu_cases_demographics_regional_view order by year desc")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3], region=r[4],
            report_type='cvsu', summary_for='region', summary_slug='demography').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='cvsu', summary_for='region', summary_slug='demography')
            db.session.add(s)
        db.session.commit()

    results = db.engine.execute("SELECT * FROM cc_attendance_regional_view order by year desc")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3], region=r[4],
            report_type='cc', summary_for='region', summary_slug='attendance').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='cc', summary_for='region', summary_slug='attendance')
            db.session.add(s)
        db.session.commit()

    # Load data for osc regional demography -sexualviolece
    results = db.engine.execute("SELECT * FROM osc_sexualviolence_demographics_regional_view order by year desc")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3], region=r[4],
            report_type='osc', summary_for='region', summary_slug='demography-sexual').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='osc', summary_for='region', summary_slug='demography-sexual')
            db.session.add(s)
        db.session.commit()

    # Load data for osc regional demography - physicalviolence
    results = db.engine.execute("SELECT * FROM osc_physicalviolence_demographics_regional_view order by year desc")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3], region=r[4],
            report_type='osc', summary_for='region', summary_slug='demography-physical').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='osc', summary_for='region', summary_slug='demography-physical')
            db.session.add(s)
        db.session.commit()

    # Load data for osc regional referredfrom pie-chart data
    results = db.engine.execute("SELECT * FROM osc_referredfrom_regional_view order by year desc")
    records = []
    for row in results:
        month = row['month']
        year = row['year']
        region_id = row['region_id']
        for k in results.keys():
            if k in ('month', 'year', 'region_id'):
                continue
            casetype, cases = (k, row[k])
            records.append((casetype, cases, month, year, region_id))

    print(records)
    for r in records:
        summary = SummaryCases.query.filter_by(
            casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), month=r[2], year=r[3], region=r[4],
            report_type='osc', summary_for='region', summary_slug='referredfrom').first()
        if summary:
            summary.value = r[1]
        else:
            s = SummaryCases(
                casetype=INDICATOR_NAME_MAPPING.get(r[0], r[0]), value=r[1], month=r[2], year=r[3],
                region=r[4], report_type='osc', summary_for='region', summary_slug='referredfrom')
            db.session.add(s)
        db.session.commit()


@app.cli.command("load_legacy_data")
@click.option('--filename', '-f')
@click.option('--report', '-r', default='pvsu')
def load_legacy_data(filename, report):
    click.echo("going to proccess the file: {0} for upload".format(filename))

    def get_last_month(date_time_obj):
        dt1 = date_time_obj.replace(day=1, hour=0, minute=0, second=0)
        dt2 = dt1 - datetime.timedelta(days=1)
        return dt2.strftime('%Y-%m')

    locs = Location.query.filter_by(level=3).all()
    districts = {}
    for l in locs:
        districts[l.id] = {'name': l.name, 'parent_id': l.parent_id}
    click.echo(districts)
    click.echo("======>")

    stations = PoliceStation.query.all()
    police_stations = {}
    for s in stations:
        police_stations[s.name] = {'id': s.id, 'district_id': s.district_id}

    courts = JusticeCourt.query.all()
    justice_courts = {}
    for c in courts:
        justice_courts[c.name] = {'id': c.id, 'district_id': c.district_id}

    click.echo(police_stations)
    click.echo("======>")

    wb = load_workbook(filename, read_only=True)
    for sheet in wb:
        # print sheet.title
        data = []
        headings = []
        j = 0
        for row in sheet.rows:
            if j > 0:
                # val = ['%s' % i.value for i in row]
                val = [u'' if i.value is None else str(i.value) for i in row]
                # print val
                data.append(val)
            else:
                headings = [u'' if i.value is None else str(i.value) for i in row]
            j += 1

        click.echo(headings)
        # click.echo(data[:10])
        for d in data:
            values = {}
            msisdn = ""
            if report in ('ncjf'):
                msisdn = d[2].strip()
            for idx, v in enumerate(d):
                if idx == 0:
                    if report in ('pvsu', 'diversion'):
                        reporting_month = d[0].split(' ')[0]
                        date_time_obj = datetime.datetime.strptime(reporting_month, '%Y-%m-%d')
                        month = get_last_month(date_time_obj)
                        year = month.split('-')[0]
                        print("====>", date_time_obj, month, year)
                    elif report in ('ncjf'):
                        month = d[0].strip()
                        year = month.split('-')[0]

                if idx == 1:
                    station = d[1].strip().title()
                    if report in ('pvsu', 'diversion'):
                        police_station = police_stations[station]['id']
                        district = police_stations[station]['district_id']
                        region = districts[district]['parent_id']
                        print("Station====>", police_station, " District: ", district, " Region: ", region)
                    if report in ('ncjf'):
                        court = justice_courts[station]['id']
                        district = police_stations[station]['district_id']
                        region = districts[district]['parent_id']

                if idx > 1:
                    if report in ('pvsu', 'diversion'):
                        values[headings[idx]] = d[idx]
                    elif report in ('ncjf'):
                        if idx == 2:
                            msisdn = d[idx]
                        else:
                            values[headings[idx]] = d[idx]

            if report in ('pvsu', 'diversion'):
                flow_data_obj = FlowData.query.filter_by(
                    year=year, month=month, report_type=report, region=region, district=district,
                    station=police_station).first()
            elif report in ('ncjf'):
                flow_data_obj = FlowData.query.filter_by(
                    year=year, month=month, report_type=report, region=region, district=district,
                    court=court).first()

            if flow_data_obj:
                # click.echo("+++++++++++++")
                # click.echo(flow_data_obj.values)
                # click.echo("--------------")
                new_values_obj = flow_data_obj.values.copy()
                for key, val in values.items():
                    new_values_obj[key] = val

                # click.echo(new_values_obj)
                flow_data_obj.values = new_values_obj
                if report in ('ncjf'):
                    flow_data_obj.msisdn = msisdn
                    flow_data_obj.updated = datetime.datetime.now()
                print(">>", msisdn, "<<", "XXXXXXXXX=>", flow_data_obj.id, flow_data_obj.year, flow_data_obj.msisdn)
                db.session.commit()
            else:
                if report in ('ncjf'):
                    db.session.add(FlowData(
                        region=region, district=district, court=court, month=month,
                        year=year, report_type=report, msisdn=msisdn, values=values))
                    db.session.commit()


@app.cli.command("load_legacy_data2")
@click.option('--filename', '-f')
@click.option('--report', '-r', default='cvsu')
def load_legacy_data2(filename, report):
    click.echo("going to proccess the file: {0} for upload".format(filename))

    def get_reporting_month(date_time_obj, month, yr=''):
        rpt_month = date_time_obj.month
        if int(month) > rpt_month:
            if yr:
                year = yr
            else:
                year = date_time_obj.year - 1
        else:
            year = date_time_obj.year
        return "{0}-{1:02}".format(year, int(month))

    locs = Location.query.filter_by(level=3).all()
    districts = {}
    district_tas = {}
    for l in locs:
        districts[l.name] = {'id': l.id, 'parent_id': l.parent_id}
        district_tas[l.id] = {}
    click.echo(districts)
    click.echo("======>")

    tas = TraditionalAuthority.query.all()
    # traditional_auths = {}
    for t in tas:
        district_tas[t.district_id][t.name] = t.id

    # click.echo(district_tas)
    # return  # XXX
    click.echo("======>")
    wb = load_workbook(filename, read_only=True)
    for sheet in wb:
        # print sheet.title
        data = []
        headings = []
        j = 0
        for row in sheet.rows:
            if j > 0:
                # val = ['%s' % i.value for i in row]
                val = [u'' if i.value is None else str(i.value) for i in row]
                # print val
                data.append(val)
            else:
                headings = [u'' if i.value is None else str(i.value) for i in row]
            j += 1

        click.echo(headings)
        click.echo(data[:10])

        for d in data:
            values = {}
            for idx, v in enumerate(d):
                if idx == 0:
                    pass

                if idx == 1:
                    reporting_month = d[1].split(' ')[0]
                    date_time_obj = datetime.datetime.strptime(reporting_month, '%Y-%m-%d')

                    month = get_reporting_month(date_time_obj, d[5], d[0])
                    year = month.split('-')[0]
                    print("====>", date_time_obj, month, year)

                if idx == 2:
                    district = d[2].strip()
                    district_id = districts[district]['id']
                    region_id = districts[district]['parent_id']
                    print(" District: ", district, " Region: ", region_id)
                if idx == 3:
                    ta = d[3].strip()
                    if ta in district_tas[district_id]:
                        ta_id = district_tas[district_id][ta]
                    else:
                        click.echo("Missing TA:{0}".format(ta))
                        pass
                if idx == 4:
                    if ta_id:
                        if report == 'cc':
                            cc_obj = ChildrensCorner.query.filter_by(
                                name=d[4].strip().title(), district_id=district_id, ta_id=ta_id).first()
                            if not cc_obj:
                                cc_obj = ChildrensCorner(
                                    name=d[4].strip().title(), district_id=district_id, ta_id=ta_id)
                                db.session.add(cc_obj)
                            cc_id = cc_obj.id

                        elif report == 'cvsu':
                            cvsu_obj = CommunityVictimSupportUnit.query.filter_by(
                                name=d[4].strip().title(), district_id=district_id, ta_id=ta_id).first()
                            if not cvsu_obj:
                                cvsu_obj = CommunityVictimSupportUnit(
                                    name=d[4].strip().title(), district_id=district_id, ta_id=ta_id)
                                db.session.add(cvsu_obj)
                            cvsu_id = cvsu_obj.id
                if idx == 5:
                    pass

                if idx > 5:
                    values[headings[idx]] = d[idx]

            click.echo(values)
            # print("++++++++++++> CVSU:", cvsu_id, " +++++++++> TA:", ta_id)

            if report == 'cvsu':
                flow_data_obj = FlowData.query.filter_by(
                    year=year, month=month, report_type=report, region=region_id, district=district_id,
                    cvsu=cvsu_id).first()
            elif report == 'cc':
                flow_data_obj = FlowData.query.filter_by(
                    year=year, month=month, report_type=report, region=region_id, district=district_id,
                    children_corner=cc_id).first()
            if flow_data_obj:
                # click.echo("+++++++++++++")
                # click.echo(flow_data_obj.values)
                # click.echo("--------------")
                new_values_obj = flow_data_obj.values.copy()
                for key, val in values.items():
                    new_values_obj[key] = val

                # click.echo(new_values_obj)
                flow_data_obj.values = new_values_obj
                print("XXXXXXXXX=>", flow_data_obj.id, flow_data_obj.year)
                db.session.commit()


@app.cli.command("create-cvsus")
@click.option('--filename', '-f')
def create_cvsu(filename):
    locs = Location.query.filter_by(level=3).all()
    districts = {}
    for l in locs:
        districts[l.name] = {'id': l.id, 'parent_id': l.parent_id}

    tas = TraditionalAuthority.query.all()
    district_tas = {}
    for t in tas:
        if t.district_id not in district_tas:
            district_tas[t.district_id] = {}
            district_tas[t.district_id][t.name] = t.id
        else:
            district_tas[t.district_id][t.name] = t.id

    print(district_tas)
    with open('../ccs2.txt') as f:
        for l in f.readlines():
            district, ta = l.strip().split(',')
            district_id = districts[district]['id']
            if ta not in district_tas[district_id]:
                # add ta
                ta_obj = TraditionalAuthority(name=ta, district_id=district_id)
                db.session.add(ta_obj)
        db.session.commit()

    with open('../uniq_cvsus.txt') as f:
        # print(f.read())
        click.echo("Gonna create CVSUs")
        for l in f.readlines():
            try:
                district, ta, cvsu = l.strip().split(',')
                district_id = districts[district]['id']
                ta_id = district_tas[district_id][ta]
                print(district_id, "=>", ta_id)
                if district_id and ta_id:
                    cvsu_obj = CommunityVictimSupportUnit.query.filter_by(
                        district_id=district_id, ta_id=ta_id, name=cvsu).first()
                    if not cvsu_obj:
                        cvsu_obj = CommunityVictimSupportUnit(
                            district_id=district_id, ta_id=ta_id, name=cvsu)
                        db.session.add(cvsu_obj)
            except:
                pass
            db.session.commit()

    with open('../uniq_ccs.txt') as f:
        # print(f.read())
        click.echo("Gonna create Children's Corners")
        for l in f.readlines():
            try:
                district, ta, cc = l.strip().split(',')
                district_id = districts[district]['id']
                ta_id = district_tas[district_id][ta]
                print(district_id, "=>", ta_id)
                if district_id and ta_id:
                    cc_obj = ChildrensCorner.query.filter_by(
                        district_id=district_id, ta_id=ta_id, name=cc).first()
                    if not cc_obj:
                        cc_obj = ChildrensCorner(
                            district_id=district_id, ta_id=ta_id, name=cc)
                        db.session.add(cc_obj)
            except:
                pass
            db.session.commit()
        click.echo("Done creating CVSUs & CCS")
